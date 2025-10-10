"""
Export functionality for CSV and Excel files
"""

import csv
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional

from .time_estimator import TimeEstimator
from .file_operations import FileOperations


class Exporters:
    """Handles exporting assignment data to various formats"""
    
    @staticmethod
    def export_csv(meta_list: List[Dict[str, Any]], 
                  csv_path: Path, 
                  summary: Optional[Dict[str, Any]] = None) -> None:
        """Export assignment metadata to CSV. 
        
        Columns: name,hash,id_item,person_label,seconds,readable,dst_path,group_name,folder_order
        If summary is provided, append a summary section at the end with per-group totals.
        """
        csv_path = Path(csv_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["name", "hash", "id_item", "person_label", "seconds", "readable", "dst_path", "group_name", "folder_order"])
            
            for m in meta_list:
                name = m.get("name")
                h = m.get("hash")
                id_item = m.get("id_item")
                label = m.get("person_label")
                seconds = m.get("seconds", 0.0)
                readable = TimeEstimator.human_readable(seconds)
                dstp = str(m.get("dst_path")) if m.get("dst_path") is not None else ""
                group_name = Path(dstp).parent.name if dstp else ""
                # Use folder_order from metadata (per-person numbering)
                folder_num = m.get("folder_order", 0)
                writer.writerow([name, h, id_item, label, seconds, readable, dstp, group_name, folder_num])

            # Append summary if provided
            if summary is not None:
                writer.writerow([])
                writer.writerow(["Summary"])
                writer.writerow(["group_label", "file_count", "total_seconds", "total_seconds_readable", 
                               "adjusted_seconds", "adjusted_readable", "unique_id_items", "unique_hashes"])
                
                total_files = 0
                total_seconds = 0.0
                total_adjusted = 0.0
                total_unique_ids = 0
                total_unique_hashes = 0
                
                for label, data in summary.items():
                    fc = data.get("file_count", 0)
                    secs = data.get("total_seconds", 0.0)
                    adj = data.get("adjusted_seconds", 0.0)
                    uid = data.get("unique_id_items", 0)
                    uhash = data.get("unique_hashes", 0)
                    writer.writerow([label, fc, secs, TimeEstimator.human_readable(secs), 
                                   adj, TimeEstimator.human_readable(adj), uid, uhash])
                    total_files += fc
                    total_seconds += secs
                    total_adjusted += adj
                    total_unique_ids += uid
                    total_unique_hashes += uhash
                
                writer.writerow([])
                writer.writerow(["TOTAL", total_files, total_seconds, TimeEstimator.human_readable(total_seconds), 
                               total_adjusted, TimeEstimator.human_readable(total_adjusted), total_unique_ids, total_unique_hashes])

        print(f"CSV exported to: {csv_path}")

    @staticmethod
    def export_xlsx(meta_list: List[Dict[str, Any]], 
                   xlsx_path: Path, 
                   summary: Optional[Dict[str, Any]] = None, 
                   person_labels: Optional[List[str]] = None) -> None:
        """Export assignment to an Excel workbook with one sheet per person and a Summary sheet.
        
        Requires openpyxl. If it's not installed, prints an instructional message and returns.
        """
        try:
            from openpyxl import Workbook
        except Exception:
            print("openpyxl is required to write .xlsx files. Install with: pip install openpyxl", file=sys.stderr)
            return

        if person_labels is None:
            person_labels = ["A", "B", "C"]

        wb = Workbook()
        # remove the default sheet created by Workbook()
        if wb.active is not None:
            wb.remove(wb.active)

        headers = ["name", "hash", "id_item", "person_label", "seconds", "readable", "dst_path", "group_name", "folder_order"]
        
        # create one sheet per person and populate rows
        for label in person_labels:
            ws = wb.create_sheet(title=label)
            ws.append(headers)
            for m in meta_list:
                if m.get("person_label") != label:
                    continue
                name = m.get("name")
                h = m.get("hash")
                id_item = m.get("id_item")
                pl = m.get("person_label")
                seconds = m.get("seconds", 0.0)
                readable = TimeEstimator.human_readable(seconds)
                dstp = str(m.get("dst_path")) if m.get("dst_path") is not None else ""
                group_name = Path(dstp).parent.name if dstp else ""
                # Use folder_order from metadata (per-person numbering)
                folder_num = m.get("folder_order", 0)
                ws.append([name, h, id_item, pl, seconds, readable, dstp, group_name, folder_num])

        # Summary sheet
        ws = wb.create_sheet(title="Summary")
        ws.append(["group_label", "file_count", "total_seconds", "total_seconds_readable", 
                  "adjusted_seconds", "adjusted_readable", "unique_id_items", "unique_hashes"])
        
        if summary is not None:
            total_files = 0
            total_secs = 0.0
            total_adj = 0.0
            total_unique_ids = 0
            total_unique_hashes = 0
            
            for label, data in summary.items():
                fc = data.get("file_count", 0)
                secs = data.get("total_seconds", 0.0)
                adj = data.get("adjusted_seconds", 0.0)
                uid = data.get("unique_id_items", 0)
                uhash = data.get("unique_hashes", 0)
                ws.append([label, fc, secs, TimeEstimator.human_readable(secs), 
                          adj, TimeEstimator.human_readable(adj), uid, uhash])
                total_files += fc
                total_secs += secs
                total_adj += adj
                total_unique_ids += uid
                total_unique_hashes += uhash
            
            ws.append([])
            ws.append(["TOTAL", total_files, total_secs, TimeEstimator.human_readable(total_secs), 
                      total_adj, TimeEstimator.human_readable(total_adj), total_unique_ids, total_unique_hashes])

        xlsx_path = Path(xlsx_path)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(xlsx_path))
        print(f"XLSX exported to: {xlsx_path}")